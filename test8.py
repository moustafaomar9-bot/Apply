# =============================================================================
#  APPLY — Sales Management Portal  |  streamlit run app.py
# =============================================================================
#
#  ┌──────────────────────────────────────────────────────────────────────────┐
#  │  DEFAULT CREDENTIALS                                                     │
#  │                                                                          │
#  │  ADMIN                                                                   │
#  │    Email    : admin@apply.com                                            │
#  │    Password : Apply@Admin2026                                            │
#  │                                                                          │
#  │  AGENTS  (auto-created when Excel is uploaded)                           │
#  │    Email    : agent<CODE>@apply.com   →  e.g.  agent250712@apply.com    │
#  │    Password : Apply<CODE>             →  e.g.  Apply250712              │
#  │                                                                          │
#  │  Change any password via  Admin → Users tab.                             │
#  └──────────────────────────────────────────────────────────────────────────┘
#
#  RUNTIME FILES  (persist across restarts — do NOT delete)
#    apply_data/data.json      – all customer records + feedback
#    apply_data/users.json     – user accounts & hashed passwords
#    apply_data/last_df.xlsx   – last uploaded client sheet
#
#  REQUIREMENTS
#    pip install streamlit pandas openpyxl
# =============================================================================

import streamlit as st
import pandas as pd
import json, os, hashlib, io
from datetime import datetime, date
from collections import Counter

# ══════════════════════════════════════════════════════════════════════
#  PATHS
# ══════════════════════════════════════════════════════════════════════
DATA_DIR   = "apply_data"
DATA_FILE  = os.path.join(DATA_DIR, "data.json")
USERS_FILE = os.path.join(DATA_DIR, "users.json")
LAST_DF    = os.path.join(DATA_DIR, "last_df.xlsx")
EXCEL_SEED = "Final_Apply_Feedback.xlsx"
os.makedirs(DATA_DIR, exist_ok=True)

# ══════════════════════════════════════════════════════════════════════
#  CONSTANTS
# ══════════════════════════════════════════════════════════════════════
ADMIN_EMAIL = "admin@apply.com"
ADMIN_PASS  = "Apply@Admin2026"

FEEDBACK_OPTIONS = [
    "done", "recall", "N.A", "closed", "travel",
    "out of area", "member", "not interested",
    "Member With other Phone",
]

FB_COLORS = {
    "done":                    "#15803d",
    "recall":                  "#b45309",
    "N.A":                     "#64748b",
    "closed":                  "#b91c1c",
    "travel":                  "#1d4ed8",
    "out of area":             "#6d28d9",
    "member":                  "#0e7490",
    "not interested":          "#475569",
    "Member With other Phone": "#5b21b6",
}
FB_BG = {
    "done":                    "#dcfce7",
    "recall":                  "#fef3c7",
    "N.A":                     "#f1f5f9",
    "closed":                  "#fee2e2",
    "travel":                  "#dbeafe",
    "out of area":             "#ede9fe",
    "member":                  "#cffafe",
    "not interested":          "#f1f5f9",
    "Member With other Phone": "#ede9fe",
}

DS_OPTIONS = ["old member", "recommendation", "Facebook", "Instagram", "other"]
DS_COLORS  = {
    "old member":     "#0e7490",
    "recommendation": "#15803d",
    "Facebook":       "#1d4ed8",
    "Instagram":      "#be185d",
    "other":          "#475569",
}
DS_BG = {
    "old member":     "#cffafe",
    "recommendation": "#dcfce7",
    "Facebook":       "#dbeafe",
    "Instagram":      "#fce7f3",
    "other":          "#f1f5f9",
}
DS_NEEDS_EXTRA = {"recommendation", "Facebook", "Instagram", "other"}

MONTHS_AR = {
    1:"يناير", 2:"فبراير", 3:"مارس", 4:"أبريل",
    5:"مايو",  6:"يونيو",  7:"يوليو", 8:"أغسطس",
    9:"سبتمبر",10:"أكتوبر",11:"نوفمبر",12:"ديسمبر",
}

# ══════════════════════════════════════════════════════════════════════
#  PAGE CONFIG + CSS
# ══════════════════════════════════════════════════════════════════════
st.set_page_config(page_title="Apply", page_icon="📋", layout="wide")

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Syne:wght@400;600;700;800;900&family=DM+Sans:wght@300;400;500;600&display=swap');

*, *::before, *::after { font-family: 'DM Sans', sans-serif !important; box-sizing: border-box; }

html, body, .stApp { background: #f6f8fc !important; color: #1e2d45 !important; font-size: 15px !important; }
.stApp > header { background: transparent !important; }

section[data-testid="stSidebar"] {
    background: #ffffff !important; border-right: 1px solid #e4eaf3 !important;
    box-shadow: 2px 0 10px rgba(0,0,0,.05); }
section[data-testid="stSidebar"] * { color: #334155 !important; font-size: 14px !important; }

h1 { color: #1e40af !important; font-family: 'Syne', sans-serif !important; font-weight: 900 !important; font-size: 1.9rem !important; }
h2 { color: #1e40af !important; font-family: 'Syne', sans-serif !important; font-weight: 800 !important; }
h3 { color: #2563eb !important; font-weight: 700 !important; }

.stTextInput input, .stTextArea textarea, .stNumberInput input {
    background: #ffffff !important; border: 1.5px solid #d0dae8 !important;
    border-radius: 10px !important; color: #1e2d45 !important; font-size: 14px !important; }
.stTextInput input:focus, .stTextArea textarea:focus {
    border-color: #3b82f6 !important; box-shadow: 0 0 0 3px rgba(59,130,246,.12) !important; }
.stTextInput input::placeholder, .stTextArea textarea::placeholder { color: #a0aec0 !important; }

div[data-baseweb="select"] > div {
    background: #ffffff !important; border: 1.5px solid #d0dae8 !important;
    border-radius: 10px !important; color: #1e2d45 !important; font-size: 14px !important; }
div[data-baseweb="select"] span, div[data-baseweb="select"] div { color: #1e2d45 !important; font-size: 14px !important; }
div[data-baseweb="popover"], div[data-baseweb="menu"], ul[data-baseweb="menu"] {
    background: #ffffff !important; border: 1px solid #e4eaf3 !important;
    border-radius: 12px !important; box-shadow: 0 8px 28px rgba(0,0,0,.10) !important; }
li[role="option"], div[role="option"] { background: #ffffff !important; color: #1e2d45 !important; font-size: 14px !important; }
li[role="option"]:hover, div[role="option"]:hover,
li[aria-selected="true"], div[aria-selected="true"] { background: #eff6ff !important; color: #1d4ed8 !important; }

.stTextInput > label, .stSelectbox > label, .stTextArea > label,
.stFileUploader > label, label { color: #64748b !important; font-weight: 600 !important; font-size: 13px !important; }

.stButton > button, [data-testid="stFormSubmitButton"] > button {
    background: linear-gradient(135deg, #1d4ed8, #2563eb) !important;
    color: #fff !important; border: none !important; border-radius: 10px !important;
    font-weight: 700 !important; font-size: 14px !important;
    box-shadow: 0 2px 8px rgba(37,99,235,.25) !important; }
.stButton > button:hover, [data-testid="stFormSubmitButton"] > button:hover { opacity: .88 !important; }
.stDownloadButton > button {
    background: linear-gradient(135deg, #065f46, #059669) !important;
    color: #fff !important; border: none !important; border-radius: 10px !important;
    font-weight: 700 !important; font-size: 14px !important;
    box-shadow: 0 2px 8px rgba(5,150,105,.25) !important; }

.stTabs [data-baseweb="tab-list"] {
    background: #eaf0f8 !important; border-radius: 12px; padding: 4px; gap: 4px; border: none !important; }
.stTabs [data-baseweb="tab"] { color: #64748b !important; font-weight: 600; border-radius: 9px; font-size: 14px !important; }
.stTabs [aria-selected="true"] {
    background: #ffffff !important; color: #1d4ed8 !important;
    box-shadow: 0 2px 8px rgba(0,0,0,.08) !important; font-weight: 700 !important; }

details > summary, .streamlit-expanderHeader {
    background: #ffffff !important; border: 1px solid #e4eaf3 !important;
    border-radius: 10px !important; color: #1e2d45 !important; font-weight: 700 !important;
    font-size: 14px !important; box-shadow: 0 1px 4px rgba(0,0,0,.05) !important; }
.streamlit-expanderContent {
    background: #f9fbfd !important; border: 1px solid #e4eaf3 !important;
    border-top: none !important; border-radius: 0 0 10px 10px !important; }

.stSuccess > div { background: #f0fdf4 !important; color: #166534 !important; border: 1px solid #bbf7d0 !important; border-radius: 10px !important; font-size: 14px !important; }
.stError   > div { background: #fef2f2 !important; color: #991b1b !important; border: 1px solid #fecaca !important; border-radius: 10px !important; font-size: 14px !important; }
.stWarning > div { background: #fffbeb !important; color: #92400e !important; border: 1px solid #fde68a !important; border-radius: 10px !important; font-size: 14px !important; }
.stInfo    > div { background: #eff6ff !important; color: #1e40af !important; border: 1px solid #bfdbfe !important; border-radius: 10px !important; font-size: 14px !important; }

.stCaption, [data-testid="stCaptionContainer"] { color: #64748b !important; font-size: 13px !important; }

[data-testid="metric-container"] {
    background: #ffffff !important; border: 1px solid #e4eaf3 !important;
    border-radius: 14px !important; padding: 18px 20px !important;
    box-shadow: 0 1px 6px rgba(0,0,0,.05) !important; }
[data-testid="stMetricLabel"]  { color: #64748b !important; font-size: 12px !important; text-transform: uppercase; letter-spacing: .8px; font-weight: 600 !important; }
[data-testid="stMetricValue"]  { color: #1e2d45 !important; font-family: 'Syne', sans-serif !important; font-size: 2.2rem !important; font-weight: 800 !important; }
[data-testid="stMetricDelta"]  { color: #059669 !important; font-size: 14px !important; font-weight: 700 !important; }

.stDataFrame iframe { border-radius: 12px; border: 1px solid #e4eaf3; }

hr { border-color: #e4eaf3 !important; margin: 18px 0 !important; }
::-webkit-scrollbar { width: 5px; height: 5px; }
::-webkit-scrollbar-track { background: #f1f5f9; }
::-webkit-scrollbar-thumb { background: #cbd5e1; border-radius: 4px; }

/* ── layout ── */
.ap-hero {
    background: linear-gradient(135deg, #ffffff 0%, #f0f5ff 100%);
    border: 1px solid #dbeafe; border-radius: 18px;
    padding: 26px 30px; margin-bottom: 22px;
    box-shadow: 0 2px 14px rgba(37,99,235,.07); }
.ap-hero-title { font-family: 'Syne', sans-serif; font-size: 1.6rem; font-weight: 900; color: #1e2d45; }
.ap-hero-sub   { color: #64748b; font-size: 14px; margin-top: 5px; }

.m-grid { display: grid; grid-template-columns: repeat(auto-fit, minmax(110px, 1fr)); gap: 10px; margin-top: 16px; }
.m-card { background: #f8fafc; border: 1px solid #e4eaf3; border-radius: 12px; padding: 14px 8px; text-align: center; }
.m-num  { font-family: 'Syne', sans-serif; font-size: 2rem; font-weight: 900; line-height: 1; }
.m-lbl  { font-size: 11px; color: #94a3b8; margin-top: 6px; text-transform: uppercase; letter-spacing: .5px; }

.kpi-grid { display: grid; grid-template-columns: repeat(auto-fit, minmax(200px, 1fr)); gap: 14px; margin: 14px 0; }
.kpi-card { background: #ffffff; border: 1px solid #e4eaf3; border-radius: 14px; padding: 20px 22px; box-shadow: 0 1px 6px rgba(0,0,0,.05); }
.kpi-ttl  { font-size: 11px; color: #94a3b8; font-weight: 700; margin-bottom: 8px; text-transform: uppercase; letter-spacing: .8px; }
.kpi-val  { font-family: 'Syne', sans-serif; font-size: 2.4rem; font-weight: 900; line-height: 1.1; }
.kpi-sub  { font-size: 14px; color: #64748b; margin-top: 6px; }
.kpi-bar-wrap { background: #f1f5f9; border-radius: 5px; height: 7px; margin-top: 10px; overflow: hidden; }
.kpi-bar      { height: 7px; border-radius: 5px; }

.sec-title {
    font-family: 'Syne', sans-serif; font-size: 1.05rem; font-weight: 800; color: #1e40af;
    border-bottom: 2px solid #dbeafe; padding-bottom: 9px; margin: 24px 0 16px; }

.client-card {
    background: #ffffff; border: 1px solid #e4eaf3;
    border-radius: 12px; padding: 15px 20px; margin-bottom: 9px;
    box-shadow: 0 1px 4px rgba(0,0,0,.04); }
.c-name { font-size: 15px; font-weight: 700; color: #1e2d45; }
.c-meta { font-size: 13px; color: #94a3b8; margin-top: 5px; }

.badge { display: inline-block; border-radius: 6px; padding: 3px 11px; font-size: 12px; font-weight: 700; margin-right: 4px; }

.info-box {
    background: #eff6ff; border: 1px solid #bfdbfe; border-left: 3px solid #3b82f6;
    border-radius: 10px; padding: 14px 18px; margin-bottom: 14px;
    font-size: 14px; color: #1e40af; line-height: 1.8; }

.stat-cards { display: grid; grid-template-columns: repeat(auto-fit, minmax(135px, 1fr)); gap: 12px; margin: 14px 0; }
.stat-card  { background: #ffffff; border: 1px solid #e4eaf3; border-radius: 12px;
               padding: 18px 12px; text-align: center; box-shadow: 0 1px 4px rgba(0,0,0,.04); }
.stat-num   { font-family: 'Syne', sans-serif; font-size: 2rem; font-weight: 900; line-height: 1; }
.stat-lbl   { font-size: 11px; color: #64748b; margin-top: 6px; text-transform: uppercase;
               letter-spacing: .5px; font-weight: 600; }

.extra-box { background: #f0f7ff; border: 1.5px solid #bfdbfe; border-radius: 10px; padding: 12px 16px; margin-top: 8px; }

.login-wrap { max-width: 420px; margin: 50px auto; }
.login-box  { background: #ffffff; border: 1px solid #e4eaf3; border-radius: 20px;
               padding: 38px 34px; box-shadow: 0 4px 24px rgba(0,0,0,.08); }
</style>
""", unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════════
#  HELPERS
# ══════════════════════════════════════════════════════════════════════
def hash_pw(pw: str) -> str:
    return hashlib.sha256(pw.encode()).hexdigest()

def pct(a, b):
    return round(a / b * 100) if b else 0

def bar_html(val, total, color):
    p = pct(val, total)
    return (f'<div class="kpi-bar-wrap">'
            f'<div class="kpi-bar" style="width:{p}%;background:{color}"></div></div>')

def fb_badge(fb: str) -> str:
    fb = str(fb).strip()
    fg = FB_COLORS.get(fb, "#475569")
    bg = FB_BG.get(fb, "#f1f5f9")
    return f'<span class="badge" style="background:{bg};color:{fg};border:1px solid {fg}30">{fb}</span>'

def ds_badge(ds: str) -> str:
    ds = str(ds).strip()
    fg = DS_COLORS.get(ds, "#475569")
    bg = DS_BG.get(ds, "#f1f5f9")
    return f'<span class="badge" style="background:{bg};color:{fg};border:1px solid {fg}30">{ds}</span>'

def fix_mobile(x) -> str:
    """Ensure mobile number is 11 digits with leading zero."""
    if x is None or str(x).strip() in ("", "None", "nan"):
        return ""
    try:
        n = str(int(float(str(x).strip())))
        return n.zfill(11)
    except Exception:
        return str(x).strip()

# ══════════════════════════════════════════════════════════════════════
#  PERSISTENCE
# ══════════════════════════════════════════════════════════════════════
def _load_users() -> dict:
    if os.path.exists(USERS_FILE):
        with open(USERS_FILE, encoding="utf-8") as f:
            return json.load(f)
    default = {ADMIN_EMAIL: {"password": hash_pw(ADMIN_PASS),
                              "role": "admin", "agent_code": None, "name": "Admin"}}
    _save_users(default)
    return default

def _save_users(u: dict):
    with open(USERS_FILE, "w", encoding="utf-8") as f:
        json.dump(u, f, ensure_ascii=False, indent=2)

def _load_records() -> list:
    if os.path.exists(DATA_FILE):
        with open(DATA_FILE, encoding="utf-8") as f:
            return json.load(f)
    return []

def _save_records(recs: list):
    with open(DATA_FILE, "w", encoding="utf-8") as f:
        json.dump(recs, f, ensure_ascii=False, indent=2)

def _migrate_feedback(records: list) -> int:
    """Rename legacy feedback values in-place. Returns count changed."""
    _map = {"Refused": "not interested", "child": "not interested"}
    n = 0
    for r in records:
        fb = str(r.get("Feedback (Sales)","")).strip()
        if fb in _map:
            r["Feedback (Sales)"] = _map[fb]
            n += 1
    return n

def _normalise_df(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df.columns = [str(c).strip() for c in df.columns]
    df = df.where(pd.notna(df), None)
    MIGRATE = {"Refused": "not interested", "child": "not interested"}
    for col, fn in [
        ("Assign Data",          lambda x: str(x)[:10] if x else ""),
        ("Agent Code",           lambda x: str(int(float(x))) if x is not None else ""),
        ("Feedback (Sales)",     lambda x: MIGRATE.get(str(x).strip(), str(x).strip()) if x else ""),
        ("Mobile",               fix_mobile),
        ("Data Source Feedback", lambda x: str(x).strip() if x else ""),
    ]:
        if col in df.columns:
            df[col] = df[col].apply(fn)
    return df

def _to_excel_bytes(records: list) -> bytes:
    """Export to Excel keeping Mobile as text with leading zero."""
    from openpyxl.styles import numbers as xl_numbers
    df  = pd.DataFrame(records)
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as w:
        df.to_excel(w, index=False, sheet_name="Apply")
        ws = w.sheets["Apply"]
        # find Mobile column index (1-based in openpyxl)
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column+1)]
        if "Mobile" in headers:
            col_idx = headers.index("Mobile") + 1
            for row in range(2, ws.max_row+1):
                cell = ws.cell(row, col_idx)
                val  = str(cell.value) if cell.value is not None else ""
                if val and val not in ("None","nan",""):
                    # pad to 11 digits
                    try: val = str(int(float(val))).zfill(11)
                    except Exception: pass
                    cell.value          = val
                    cell.number_format  = "@"   # force text format in Excel
    return buf.getvalue()

def _save_df(df: pd.DataFrame):
    """Save last uploaded df, keeping Mobile as text."""
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import numbers as xl_numbers
        df = df.copy()
        if "Mobile" in df.columns:
            df["Mobile"] = df["Mobile"].apply(
                lambda x: str(int(float(str(x)))).zfill(11)
                if str(x).strip() not in ("","None","nan") else "")
        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine="openpyxl") as w:
            df.to_excel(w, index=False, sheet_name="Apply")
            ws = w.sheets["Apply"]
            headers = [ws.cell(1,c).value for c in range(1, ws.max_column+1)]
            if "Mobile" in headers:
                ci = headers.index("Mobile") + 1
                for row in range(2, ws.max_row+1):
                    ws.cell(row, ci).number_format = "@"
        with open(LAST_DF, "wb") as f:
            f.write(buf.getvalue())
    except Exception:
        pass

# ══════════════════════════════════════════════════════════════════════
#  AGENT ACCOUNTS
# ══════════════════════════════════════════════════════════════════════
def _auto_create_agents(records: list) -> list:
    users = _load_users(); created = []
    for r in records:
        ac = str(r.get("Agent Code","")).strip()
        if not ac: continue
        email = f"agent{ac}@apply.com"
        if email not in users:
            users[email] = {"password": hash_pw(f"Apply{ac}"),
                            "role": "sales", "agent_code": ac, "name": f"Agent {ac}"}
            created.append(email)
    if created: _save_users(users)
    return created

# ══════════════════════════════════════════════════════════════════════
#  STATS
# ══════════════════════════════════════════════════════════════════════
def _get_stats(records: list) -> dict:
    total  = len(records)
    fb_cnt = Counter(str(r.get("Feedback (Sales)","")).strip() for r in records)
    ds_cnt = Counter(str(r.get("Data Source Feedback","")).strip() for r in records)
    return {
        "total":  total,
        "fb":     dict(fb_cnt),
        "ds":     dict(ds_cnt),
        "done":   fb_cnt.get("done", 0),
        "recall": fb_cnt.get("recall", 0),
        "closed": fb_cnt.get("closed", 0),
        "ni":     fb_cnt.get("not interested", 0),
        "na":     fb_cnt.get("N.A", 0),
    }

# ══════════════════════════════════════════════════════════════════════
#  SESSION INIT
# ══════════════════════════════════════════════════════════════════════
for k, v in [("logged_in", False), ("role", None),
              ("agent_code", None), ("user_name", ""), ("user_email", "")]:
    if k not in st.session_state:
        st.session_state[k] = v

# ══════════════════════════════════════════════════════════════════════
#  LOGIN
# ══════════════════════════════════════════════════════════════════════
if not st.session_state.logged_in:
    st.markdown('<div class="login-wrap">', unsafe_allow_html=True)
    st.markdown("""
    <div class="login-box">
      <div style="text-align:center;margin-bottom:24px">
        <div style="font-family:'Syne',sans-serif;font-size:2.6rem;font-weight:900;
                    background:linear-gradient(135deg,#3b82f6,#06b6d4);
                    -webkit-background-clip:text;-webkit-text-fill-color:transparent;
                    letter-spacing:5px">APPLY</div>
        <div style="color:#94a3b8;font-size:.9rem;letter-spacing:1px;margin-top:6px">
          Sales Management Portal
        </div>
      </div>
    </div>""", unsafe_allow_html=True)
    email    = st.text_input("📧 Email", placeholder="agent250712@apply.com")
    password = st.text_input("🔒 Password", type="password", placeholder="••••••••")
    st.markdown("<br>", unsafe_allow_html=True)
    if st.button("Sign In →", use_container_width=True):
        e = email.strip().lower(); users = _load_users()
        if e == ADMIN_EMAIL and password == ADMIN_PASS:
            st.session_state.update(logged_in=True, role="admin", user_email=e, user_name="Admin")
            st.rerun()
        elif e in users and users[e]["password"] == hash_pw(password):
            u = users[e]
            st.session_state.update(logged_in=True, role=u["role"], user_email=e,
                                    agent_code=u.get("agent_code"), user_name=u["name"])
            st.rerun()
        else:
            st.error("❌ Incorrect email or password.")
    st.markdown("</div>", unsafe_allow_html=True)
    st.stop()

# auto-migrate legacy feedback values on every startup
if os.path.exists(DATA_FILE):
    _mr = _load_records()
    _mn = _migrate_feedback(_mr)
    if _mn > 0:
        _save_records(_mr)

# ══════════════════════════════════════════════════════════════════════
#  SEED on first run
# ══════════════════════════════════════════════════════════════════════
if not os.path.exists(DATA_FILE) and os.path.exists(EXCEL_SEED):
    _seed = _normalise_df(pd.read_excel(EXCEL_SEED))
    _recs = _seed.to_dict(orient="records")
    _save_records(_recs)
    _auto_create_agents(_recs)
    _save_df(_seed)

# ══════════════════════════════════════════════════════════════════════
#  SIDEBAR
# ══════════════════════════════════════════════════════════════════════
with st.sidebar:
    st.markdown(f"""
    <div style="padding:18px 4px 12px;text-align:center">
      <div style="font-family:'Syne',sans-serif;font-size:1.65rem;font-weight:900;
                  background:linear-gradient(135deg,#3b82f6,#06b6d4);
                  -webkit-background-clip:text;-webkit-text-fill-color:transparent;
                  letter-spacing:4px">APPLY</div>
    </div>
    <div style="background:#f0f5ff;border:1px solid #dbeafe;border-radius:12px;
                padding:12px 14px;margin-bottom:14px">
      <div style="font-weight:700;color:#1e2d45;font-size:.95rem">
        {st.session_state.user_name}</div>
      <div style="color:#64748b;font-size:.78rem;margin-top:2px">
        {'🛡️ Administrator' if st.session_state.role == 'admin'
         else f'👤 Agent · {st.session_state.agent_code}'}
      </div>
    </div>""", unsafe_allow_html=True)

    pages = (["Dashboard", "All Data", "Upload Data", "Users", "Monthly Report"]
             if st.session_state.role == "admin"
             else ["My Dashboard", "My Clients"])
    page = st.radio("nav", pages, label_visibility="collapsed")

    if st.session_state.role == "admin":
        st.markdown("---")
        st.markdown("**🔑 Agent Credentials**")
        _all_u = _load_users()
        _ags   = {e: v for e, v in _all_u.items() if v.get("role") == "sales"}
        if _ags:
            for _em, _info in _ags.items():
                st.markdown(f"`{_em}`  \n`Apply{_info['agent_code']}`")
        else:
            st.caption("Upload a sheet to generate accounts.")

    st.markdown("---")
    if st.button("Sign Out", use_container_width=True):
        for k in ["logged_in","role","agent_code","user_name","user_email"]:
            st.session_state[k] = False if k == "logged_in" else None if k in ["role","agent_code"] else ""
        st.rerun()

# ══════════════════════════════════════════════════════════════════════
#  SHARED: STAT CARDS
# ══════════════════════════════════════════════════════════════════════
def _render_fb_cards(records: list):
    s = _get_stats(records); total = s["total"]
    if not total: return
    cards = ""
    for fb in FEEDBACK_OPTIONS:
        cnt   = s["fb"].get(fb, 0)
        color = FB_COLORS.get(fb, "#475569")
        bg    = FB_BG.get(fb, "#f8fafc")
        cards += (f'<div class="stat-card" style="border-left:3px solid {color};background:{bg}">'
                  f'<div class="stat-num" style="color:{color}">{cnt}</div>'
                  f'<div class="stat-lbl" style="color:{color}">{fb}</div>'
                  f'<div style="font-size:12px;color:#94a3b8;margin-top:4px">{pct(cnt,total)}%</div></div>')
    st.markdown(f'<div class="stat-cards">{cards}</div>', unsafe_allow_html=True)

def _render_ds_cards(records: list):
    s = _get_stats(records); total = s["total"]
    if not total: return
    all_ds = set(str(r.get("Data Source Feedback","")).strip() for r in records) | set(DS_OPTIONS)
    all_ds.discard(""); all_ds.discard("None"); all_ds.discard("nan")
    cards = ""
    for ds in sorted(all_ds):
        cnt   = s["ds"].get(ds, 0)
        if cnt == 0: continue
        color = DS_COLORS.get(ds, "#475569")
        bg    = DS_BG.get(ds, "#f8fafc")
        cards += (f'<div class="stat-card" style="border-left:3px solid {color};background:{bg}">'
                  f'<div class="stat-num" style="color:{color}">{cnt}</div>'
                  f'<div class="stat-lbl" style="color:{color}">{ds}</div>'
                  f'<div style="font-size:12px;color:#94a3b8;margin-top:4px">{pct(cnt,total)}%</div></div>')
    if cards:
        st.markdown(f'<div class="stat-cards">{cards}</div>', unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════════
#  PAGE: ADMIN DASHBOARD
# ══════════════════════════════════════════════════════════════════════
def page_dashboard():
    records = _load_records()
    s       = _get_stats(records)

    st.markdown(f"""
    <div class="ap-hero">
      <div class="ap-hero-title">🛡️ Admin Dashboard</div>
      <div class="ap-hero-sub">Total records: {s['total']}</div>
    </div>""", unsafe_allow_html=True)

    k1,k2 = st.columns(2)
    k1.metric("Total",          s["total"])
    k2.metric("Done ✅",        s["done"],   f"{pct(s['done'],s['total'])}%")
    if not records:
        st.info("No data yet — upload a client sheet to get started."); return

    st.markdown('<div class="sec-title">Feedback Breakdown</div>', unsafe_allow_html=True)
    _render_fb_cards(records)

    st.markdown('<div class="sec-title">Data Source Breakdown</div>', unsafe_allow_html=True)
    _render_ds_cards(records)

    st.markdown('<div class="sec-title">KPI Bars</div>', unsafe_allow_html=True)
    st.markdown(f"""
    <div class="kpi-grid">
      <div class="kpi-card">
        <div class="kpi-ttl">Done Rate</div>
        <div class="kpi-val" style="color:#15803d">{pct(s['done'],s['total'])}%</div>
        <div class="kpi-sub">{s['done']} of {s['total']}</div>
        {bar_html(s['done'],s['total'],'#15803d')}
      </div>
      <div class="kpi-card">
        <div class="kpi-ttl">Recall Rate</div>
        <div class="kpi-val" style="color:#b45309">{pct(s['recall'],s['total'])}%</div>
        <div class="kpi-sub">{s['recall']} clients</div>
        {bar_html(s['recall'],s['total'],'#b45309')}
      </div>
      <div class="kpi-card">
        <div class="kpi-ttl">Closed Rate</div>
        <div class="kpi-val" style="color:#b91c1c">{pct(s['closed'],s['total'])}%</div>
        <div class="kpi-sub">{s['closed']} clients</div>
        {bar_html(s['closed'],s['total'],'#b91c1c')}
      </div>
      <div class="kpi-card">
        <div class="kpi-ttl">Not Interested</div>
        <div class="kpi-val" style="color:#475569">{pct(s['ni'],s['total'])}%</div>
        <div class="kpi-sub">{s['ni']} clients</div>
        {bar_html(s['ni'],s['total'],'#475569')}
      </div>
    </div>""", unsafe_allow_html=True)

    st.markdown('<div class="sec-title">Agent Performance</div>', unsafe_allow_html=True)
    agents = sorted({str(r.get("Agent Code","")).strip() for r in records if r.get("Agent Code")})
    perf   = []
    for ac in agents:
        mine = [r for r in records if str(r.get("Agent Code","")).strip() == ac]
        ag_s = _get_stats(mine)
        row  = {"Agent": ac, "Total": ag_s["total"]}
        # one column per feedback option
        for fb in FEEDBACK_OPTIONS:
            row[fb] = ag_s["fb"].get(fb, 0)
        row["Conv %"] = f"{pct(ag_s['done'], ag_s['total'])}%"
        perf.append(row)
    if perf:
        perf_df = pd.DataFrame(perf).sort_values("done", ascending=False).reset_index(drop=True)
        st.dataframe(perf_df, use_container_width=True, hide_index=True, height=280)

    st.markdown('<div class="sec-title">Export</div>', unsafe_allow_html=True)
    c1, c2, _ = st.columns([1,1,2])
    with c1:
        st.download_button("⬇️ Download HTML Report",
            data=_build_html_report(records),
            file_name=f"apply_report_{datetime.now():%Y%m%d_%H%M}.html", mime="text/html")
    with c2:
        st.download_button("📊 Download Excel Dashboard",
            data=_build_excel_dashboard(records),
            file_name=f"apply_dashboard_{datetime.now():%Y%m%d_%H%M}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ══════════════════════════════════════════════════════════════════════
#  PAGE: ALL DATA
# ══════════════════════════════════════════════════════════════════════
def _dedup_records(records):
    seen = {}
    for i, r in enumerate(records):
        key = str(r.get("Mobile","")).strip().lstrip("0")
        seen[key if key else f"__nokey_{i}"] = i
    cleaned = [records[i] for i in sorted(seen.values())]
    return cleaned, len(records) - len(cleaned)


def page_all_data():
    records = _load_records()
    st.markdown('<div class="sec-title">All Client Data</div>', unsafe_allow_html=True)
    if not records: st.info("No records yet."); return

    cleaned, dup_count = _dedup_records(records)
    if dup_count > 0:
        st.warning(f"⚠️ Found **{dup_count}** duplicate(s) in the database.")
        if st.button("🧹 Remove Duplicates Now", type="primary"):
            _save_records(cleaned)
            st.success(f"✅ Removed {dup_count} duplicate(s).")
            st.rerun()
        records = cleaned

    _ac_list = sorted({str(r.get("Agent Code","")).strip() for r in records if r.get("Agent Code")})
    _ac_done = {ac: sum(1 for r in records
                        if str(r.get("Agent Code","")).strip()==ac
                        and str(r.get("Feedback (Sales)","")).strip()=="done")
                for ac in _ac_list}
    agents = ["All"] + sorted(_ac_list, key=lambda x: _ac_done.get(x,0), reverse=True)

    tab_view, tab_delete, tab_transfer = st.tabs([
        "📋  View & Filter",
        "🗑️  Delete Records",
        "🔀  Transfer Record",
    ])

    # ── VIEW ──────────────────────────────────────────────────────────
    with tab_view:
        f1,f2,f3 = st.columns(3)
        with f1: sel_a = st.selectbox("Agent", agents, key="va")
        with f2: sel_f = st.selectbox("Feedback", ["All"] + FEEDBACK_OPTIONS, key="vf")
        with f3: srch  = st.text_input("Search name / mobile", placeholder="🔍", key="vs")

        flt = records[:]
        if sel_a != "All":
            flt = [r for r in flt if str(r.get("Agent Code","")).strip() == sel_a]
        if sel_f != "All":
            flt = [r for r in flt if str(r.get("Feedback (Sales)","")).strip() == sel_f]
        if srch:
            s = srch.lower().lstrip("0")
            flt = [r for r in flt
                   if s in str(r.get("Customer Name","")).lower()
                   or s in str(r.get("Mobile","")).lstrip("0")]

        st.caption(f"Showing **{len(flt)}** of **{len(records)}** records")
        if flt:
            df_show = pd.DataFrame(flt).copy()
            if "Mobile" in df_show.columns:
                df_show["Mobile"] = df_show["Mobile"].apply(
                    lambda x: str(int(float(str(x)))).zfill(11)
                    if str(x).strip() not in ("","None","nan") else "")
            st.dataframe(df_show, use_container_width=True, hide_index=True, height=450)
        else:
            st.info("No records match.")

        st.download_button(
            "⬇️ Download Filtered Excel",
            data=_to_excel_bytes(flt),
            file_name=f"apply_{datetime.now():%Y%m%d_%H%M}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    # ── DELETE ────────────────────────────────────────────────────────
    with tab_delete:
        st.markdown(
            '<div class="info-box" style="border-left-color:#b91c1c;'
            'background:#fef2f2;color:#7f1d1d">'
            '☑️ تحديد الريكوردات '
            'من الجدول، بعدين '
            'اضغط Delete.</div>',
            unsafe_allow_html=True)

        dc1,dc2,dc3 = st.columns(3)
        with dc1: da  = st.selectbox("Agent",    agents, key="da")
        with dc2: df2 = st.selectbox("Feedback", ["All"] + FEEDBACK_OPTIONS, key="df2")
        with dc3: ds2 = st.text_input("Search",  placeholder="🔍 name / mobile", key="ds2")

        del_flt = records[:]
        if da  != "All":
            del_flt = [r for r in del_flt if str(r.get("Agent Code","")).strip() == da]
        if df2 != "All":
            del_flt = [r for r in del_flt if str(r.get("Feedback (Sales)","")).strip() == df2]
        if ds2:
            sq = ds2.lower().lstrip("0")
            del_flt = [r for r in del_flt
                       if sq in str(r.get("Customer Name","")).lower()
                       or sq in str(r.get("Mobile","")).lstrip("0")]

        if not del_flt:
            st.info("No records match.")
        else:
            df_del = pd.DataFrame(del_flt).copy()
            if "Mobile" in df_del.columns:
                df_del["Mobile"] = df_del["Mobile"].apply(
                    lambda x: str(int(float(str(x)))).zfill(11)
                    if str(x).strip() not in ("","None","nan") else "")
            disp_cols = [c for c in ["Customer Name","Mobile","Agent Code",
                                     "Feedback (Sales)","Data Source Feedback","Assign Data"]
                         if c in df_del.columns]
            df_del_show = df_del[disp_cols].copy()
            df_del_show.insert(0, "Select", False)

            st.caption(f"**{len(del_flt)}** records — tick to select then press Delete:")
            edited = st.data_editor(
                df_del_show,
                use_container_width=True,
                hide_index=True,
                height=min(420, 55 + len(del_flt) * 38),
                column_config={
                    "Select": st.column_config.CheckboxColumn(
                        "🗑️ Select", width="small")
                },
                disabled=[c for c in df_del_show.columns if c != "Select"],
                key="del_editor",
            )

            selected_mask  = edited["Select"].tolist()
            selected_count = sum(selected_mask)

            if selected_count > 0:
                st.markdown(
                    f'<span style="color:#b91c1c;font-weight:700">'
                    f'Selected: {selected_count} record(s)</span>',
                    unsafe_allow_html=True)
                if st.button(
                    f"🗑️ Delete {selected_count} Record(s)",
                    type="primary", key="do_delete"):
                    to_delete = set()
                    for i, sel in enumerate(selected_mask):
                        if sel:
                            mob = str(del_flt[i].get("Mobile","")).strip().lstrip("0")
                            ac  = str(del_flt[i].get("Agent Code","")).strip()
                            to_delete.add((mob, ac))
                    all_recs = _load_records()
                    before   = len(all_recs)
                    all_recs = [
                        r for r in all_recs
                        if (str(r.get("Mobile","")).strip().lstrip("0"),
                            str(r.get("Agent Code","")).strip()) not in to_delete
                    ]
                    _save_records(all_recs)
                    st.success(f"✅ Deleted {before - len(all_recs)} record(s).")
                    st.rerun()

    # ── TRANSFER ──────────────────────────────────────────────────────
    with tab_transfer:
        st.markdown(
            '<div class="info-box"><b>🔀 Transfer client between agents</b><br>'
            'ابحث، اختار، '
            'وحوله للـ Agent الجديد.</div>',
            unsafe_allow_html=True)

        tr_srch = st.text_input(
            "🔍 Search client", placeholder="01XXXXXXXXX", key="tr_s")
        if tr_srch:
            sq = tr_srch.lower().lstrip("0")
            matches = [r for r in records
                       if sq in str(r.get("Customer Name","")).lower()
                       or sq in str(r.get("Mobile","")).lstrip("0")]
            if not matches:
                st.warning("No clients found.")
            else:
                opts = [
                    f"#{r.get('Ser','')}  {r.get('Customer Name','')}  "
                    f"📱 {str(r.get('Mobile','')).zfill(11) if str(r.get('Mobile','')).strip() not in ('','None','nan') else ''}  "
                    f"| Agent: {r.get('Agent Code','')}  | {str(r.get('Feedback (Sales)',''))}"
                    for r in matches
                ]
                sel_i   = st.selectbox(
                    "Select client", range(len(opts)),
                    format_func=lambda i: opts[i], key="tr_rec")
                sel_rec = matches[sel_i]
                cur_ac  = str(sel_rec.get("Agent Code","")).strip()
                cur_mob = str(sel_rec.get("Mobile","")).strip()

                st.markdown(f"""
                <div style="background:#f0f5ff;border:1px solid #dbeafe;
                            border-radius:10px;padding:14px 18px;margin:12px 0">
                  <b style="color:#1e2d45">{sel_rec.get('Customer Name','')}</b>
                  &nbsp;|  📱 {cur_mob.zfill(11) if cur_mob else ''}
                  &nbsp;|  Current Agent: <b style="color:#1d4ed8">{cur_ac}</b>
                </div>""", unsafe_allow_html=True)

                other_agents = [a for a in _ac_list if a != cur_ac]
                if not other_agents:
                    st.warning("No other agents to transfer to.")
                else:
                    c1, c2 = st.columns(2)
                    with c1:
                        new_ac = st.selectbox(
                            "Transfer to Agent", other_agents, key="tr_to")
                    with c2:
                        st.markdown("<br>", unsafe_allow_html=True)
                        if st.button("🔀 Transfer Now",
                                     type="primary", use_container_width=True):
                            all_recs = _load_records()
                            for r in all_recs:
                                if (str(r.get("Agent Code","")).strip() == cur_ac and
                                        str(r.get("Mobile","")).strip().lstrip("0")
                                        == cur_mob.lstrip("0")):
                                    r["Agent Code"] = new_ac
                                    break
                            _save_records(all_recs)
                            st.success(
                                f"✅ Transferred "
                                f"**{sel_rec.get('Customer Name','')}** "
                                f"→ Agent **{new_ac}**")
                            st.rerun()

#  PAGE: UPLOAD DATA
# ══════════════════════════════════════════════════════════════════════
def page_upload():
    st.markdown('<div class="sec-title">Upload Data</div>', unsafe_allow_html=True)
    tab1, tab2, tab3 = st.tabs(["📋  Client Sheet", "💬  Feedback Sheet", "⬇️  Download"])

    with tab1:
        st.markdown("""
        <div class="info-box">
          <b>📋 Client Data Sheet</b><br>
          Upload new clients to distribute to agents.<br>
          🤖 Agent accounts auto-created per Agent Code<br>
          🔑 Email: <b>agent&lt;CODE&gt;@apply.com</b> · Password: <b>Apply&lt;CODE&gt;</b>
        </div>""", unsafe_allow_html=True)
        overwrite = st.checkbox("Replace all existing data (unchecked = append)", key="ov1")
        up1 = st.file_uploader("Choose Excel (.xlsx)", type=["xlsx","xls"], key="up1")
        if up1 is not None:
            raw = up1.read()
            try: df_new = _normalise_df(pd.read_excel(io.BytesIO(raw)))
            except Exception as e: st.error(f"Cannot read file: {e}"); return
            st.markdown('<div class="sec-title">Preview</div>', unsafe_allow_html=True)
            st.dataframe(df_new.head(10), use_container_width=True, hide_index=True)
            c1,_ = st.columns([1,3])
            with c1:
                if st.button("📥 Import Clients", type="primary", use_container_width=True):
                    new_recs = df_new.to_dict(orient="records")
                    recs     = _load_records()
                    if overwrite:
                        final = new_recs
                    else:
                        # deduplicate: skip rows whose Mobile already exists
                        existing_mobs = {str(r.get("Mobile","")).strip().lstrip("0")
                                         for r in recs if r.get("Mobile")}
                        mx = max((r.get("Ser",0) for r in recs), default=0)
                        added_new = []
                        for r in new_recs:
                            mob_key = str(r.get("Mobile","")).strip().lstrip("0")
                            if mob_key and mob_key in existing_mobs:
                                continue   # skip duplicate
                            mx += 1
                            if not r.get("Ser"): r["Ser"] = mx
                            added_new.append(r)
                            existing_mobs.add(mob_key)
                        final = recs + added_new
                        if len(new_recs) - len(added_new) > 0:
                            st.warning(f"⚠️ Skipped **{len(new_recs)-len(added_new)}** duplicate mobiles.")
                    _save_records(final); _save_df(df_new)
                    created = _auto_create_agents(final)
                    st.success(f"✅ Imported **{len(new_recs)}** records!")
                    if created:
                        codes = ", ".join(e.split("@")[0].replace("agent","") for e in created)
                        st.info(f"🆕 New accounts: **{codes}**")
                    st.rerun()

    with tab2:
        st.markdown("""
        <div class="info-box">
          <b>💬 Feedback Sheet</b><br>
          Upload feedback for agents without system access.<br>
          🔁 Merge by <b>Mobile Number</b>: Found → update · Not found → add new
        </div>""", unsafe_allow_html=True)
        up2 = st.file_uploader("Choose Feedback Excel (.xlsx)", type=["xlsx","xls"], key="up2")
        if up2 is not None:
            raw2 = up2.read()
            try: df_fb = _normalise_df(pd.read_excel(io.BytesIO(raw2)))
            except Exception as e: st.error(f"Cannot read file: {e}"); return
            st.markdown('<div class="sec-title">Preview</div>', unsafe_allow_html=True)
            st.dataframe(df_fb.head(10), use_container_width=True, hide_index=True)
            recs     = _load_records()
            mob_idx  = {str(r.get("Mobile","")).strip(): i for i,r in enumerate(recs)}
            fb_recs  = df_fb.to_dict(orient="records")
            will_upd = [r for r in fb_recs if str(r.get("Mobile","")).strip() in mob_idx]
            will_add = [r for r in fb_recs if str(r.get("Mobile","")).strip() not in mob_idx]
            c1,c2 = st.columns(2)
            c1.metric("To UPDATE", len(will_upd))
            c2.metric("To ADD (new)", len(will_add))
            c1,_ = st.columns([1,3])
            with c1:
                if st.button("🔀 Merge Feedback", type="primary", use_container_width=True):
                    recs  = _load_records()
                    midx  = {str(r.get("Mobile","")).strip(): i for i,r in enumerate(recs)}
                    mx    = max((r.get("Ser",0) for r in recs), default=0)
                    updated = added = 0
                    for r in fb_recs:
                        mob = str(r.get("Mobile","")).strip()
                        fb  = str(r.get("Feedback (Sales)","")).strip()
                        if mob in midx:
                            idx = midx[mob]
                            if fb and fb not in ("None","nan",""):
                                recs[idx]["Feedback (Sales)"] = fb
                            for fld in ("Notes","Data Source Feedback","Assign Data","DS Extra"):
                                if r.get(fld): recs[idx][fld] = r[fld]
                            updated += 1
                        else:
                            mx += 1; r["Ser"] = mx; recs.append(r); added += 1
                    _save_records(recs); _auto_create_agents(recs)
                    st.success(f"✅ Merge done! Updated: **{updated}** · Added: **{added}**")
                    st.rerun()

    with tab3:
        recs = _load_records()
        st.markdown('<div class="sec-title">Download Current Data</div>', unsafe_allow_html=True)
        if recs:
            st.caption(f"Total records: **{len(recs)}**")
            st.download_button("⬇️ Download All Data (Excel)",
                data=_to_excel_bytes(recs),
                file_name=f"apply_full_{datetime.now():%Y%m%d_%H%M}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        else: st.info("No data yet.")

# ══════════════════════════════════════════════════════════════════════
#  PAGE: USERS
# ══════════════════════════════════════════════════════════════════════
def page_users():
    users = _load_users()
    st.markdown('<div class="sec-title">User Accounts</div>', unsafe_allow_html=True)
    udf = pd.DataFrame([{"Email":e,"Name":v["name"],"Role":v["role"],
                          "Agent Code":v.get("agent_code","")} for e,v in users.items()])
    st.dataframe(udf, use_container_width=True, hide_index=True)

    t1,t2,t3 = st.tabs(["➕  Add User","🔑  Change Password","🗑️  Remove"])
    with t1:
        with st.form("fadd", clear_on_submit=True):
            c1,c2 = st.columns(2)
            with c1:
                ne=st.text_input("Email *"); nn=st.text_input("Full Name *"); npw=st.text_input("Password *",type="password")
            with c2:
                nr=st.selectbox("Role",["sales","admin"]); nac=st.text_input("Agent Code (sales only)")
            if st.form_submit_button("Add User", type="primary"):
                users=_load_users(); e=ne.strip().lower()
                if not (e and nn.strip() and npw): st.error("All fields required.")
                elif e in users: st.error("Email already exists.")
                else:
                    users[e]={"password":hash_pw(npw),"role":nr,
                               "agent_code":nac.strip() if nr=="sales" else None,"name":nn.strip()}
                    _save_users(users); st.success(f"✅ **{nn.strip()}** created."); st.rerun()
    with t2:
        with st.form("fchpw"):
            users=_load_users(); cpu=st.selectbox("User",list(users.keys())); cpw=st.text_input("New Password",type="password")
            if st.form_submit_button("Update Password"):
                if not cpw: st.error("Password cannot be empty.")
                else:
                    users=_load_users(); users[cpu]["password"]=hash_pw(cpw)
                    _save_users(users); st.success("✅ Password updated.")
    with t3:
        users=_load_users(); rem=[e for e in users if e!=ADMIN_EMAIL]
        if rem:
            with st.form("fdel"):
                du=st.selectbox("Select user to remove",rem)
                if st.form_submit_button("Remove User"):
                    users=_load_users(); del users[du]
                    _save_users(users); st.warning(f"**{du}** removed."); st.rerun()
        else: st.info("No removable users.")

# ══════════════════════════════════════════════════════════════════════
#  PAGE: MONTHLY REPORT
# ══════════════════════════════════════════════════════════════════════
def page_monthly_report():
    records = _load_records()
    st.markdown('<div class="sec-title">📅 Monthly Report</div>', unsafe_allow_html=True)

    if not records:
        st.info("No data yet."); return

    # ── parse dates
    df = pd.DataFrame(records)
    df["_date"] = pd.to_datetime(df.get("Assign Data",""), errors="coerce")
    df["_year"]  = df["_date"].dt.year
    df["_month"] = df["_date"].dt.month

    valid = df.dropna(subset=["_year","_month"])
    if valid.empty:
        st.warning("No valid dates found in Assign Data column."); return

    # ── controls
    years  = sorted(valid["_year"].dropna().unique().tolist(), reverse=True)
    months = sorted(valid["_month"].dropna().unique().tolist())

    c1,c2,c3 = st.columns(3)
    with c1:
        sel_year  = st.selectbox("Year", [int(y) for y in years])
    with c2:
        month_labels = {m: f"{MONTHS_AR[int(m)]} ({int(m)})" for m in months}
        sel_month = st.selectbox("Month", [int(m) for m in months],
                                  format_func=lambda m: month_labels.get(m, str(m)))
    with c3:
        agents = ["All"] + sorted({str(r.get("Agent Code","")).strip()
                                    for r in records if r.get("Agent Code")})
        sel_agent = st.selectbox("Agent", agents)

    # ── filter
    mask = (valid["_year"] == sel_year) & (valid["_month"] == sel_month)
    month_df = valid[mask]
    if sel_agent != "All":
        month_df = month_df[month_df["Agent Code"].astype(str).str.strip() == sel_agent]

    month_recs = month_df.drop(columns=["_date","_year","_month"]).to_dict(orient="records")
    s = _get_stats(month_recs)

    month_name = MONTHS_AR.get(sel_month, str(sel_month))
    st.markdown(f"""
    <div class="ap-hero">
      <div class="ap-hero-title">📅 {month_name} {sel_year}</div>
      <div class="ap-hero-sub">
        {f'Agent {sel_agent} · ' if sel_agent != 'All' else ''}
        {s['total']} clients assigned this month
      </div>
    </div>""", unsafe_allow_html=True)

    # ── KPI metrics
    k1,k2,k3,k4,k5 = st.columns(5)
    k1.metric("Total",           s["total"])
    k2.metric("Done ✅",         s["done"],   f"{pct(s['done'],s['total'])}%")
    k3.metric("Recall 🔁",       s["recall"])
    k4.metric("Closed 🔴",       s["closed"])
    k5.metric("Not Interested",  s["ni"])

    st.markdown('<div class="sec-title">Feedback Breakdown</div>', unsafe_allow_html=True)
    _render_fb_cards(month_recs)

    st.markdown('<div class="sec-title">Data Source Breakdown</div>', unsafe_allow_html=True)
    _render_ds_cards(month_recs)

    # ── Agent comparison within month (only if All agents)
    if sel_agent == "All":
        st.markdown('<div class="sec-title">Agent Performance This Month</div>', unsafe_allow_html=True)
        agent_list = sorted({str(r.get("Agent Code","")).strip()
                              for r in month_recs if r.get("Agent Code")})
        perf = []
        for ac in agent_list:
            mine = [r for r in month_recs if str(r.get("Agent Code","")).strip() == ac]
            ag_s = _get_stats(mine)
            perf.append({"Agent": ac, "Total": ag_s["total"], "Done": ag_s["done"],
                         "Recall": ag_s["recall"], "Closed": ag_s["closed"],
                         "Conv %": f"{pct(ag_s['done'],ag_s['total'])}%"})
        if perf:
            st.dataframe(pd.DataFrame(perf).sort_values("Done", ascending=False).reset_index(drop=True),
                         use_container_width=True, hide_index=True, height=220)

    # ── Client table
    st.markdown('<div class="sec-title">Client Details</div>', unsafe_allow_html=True)
    if month_recs:
        disp_cols = [c for c in ["Ser","Customer Name","Mobile","Feedback (Sales)",
                                   "Data Source Feedback","Agent Code","Assign Data","Notes"]
                     if c in pd.DataFrame(month_recs).columns]
        st.dataframe(pd.DataFrame(month_recs)[disp_cols],
                     use_container_width=True, hide_index=True, height=350)
    else:
        st.info("No records for this selection.")

    # ── Export
    st.markdown('<div class="sec-title">Export</div>', unsafe_allow_html=True)
    c1, c2, _ = st.columns([1,1,2])
    with c1:
        st.download_button("⬇️ Download Excel",
            data=_to_excel_bytes(month_recs),
            file_name=f"apply_{month_name}_{sel_year}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    with c2:
        label = f"Agent {sel_agent}" if sel_agent != "All" else None
        st.download_button("⬇️ Download HTML Report",
            data=_build_html_report(month_recs, agent_label=label,
                                     title=f"{month_name} {sel_year}"),
            file_name=f"apply_report_{month_name}_{sel_year}.html",
            mime="text/html")

# ══════════════════════════════════════════════════════════════════════
#  PAGE: SALES — MY DASHBOARD
# ══════════════════════════════════════════════════════════════════════
def page_my_dashboard():
    records = _load_records()
    agent   = str(st.session_state.agent_code).strip()
    mine    = [r for r in records if str(r.get("Agent Code","")).strip() == agent]
    s       = _get_stats(mine)

    st.markdown(f"""
    <div class="ap-hero" style="border-color:#dbeafe">
      <div class="ap-hero-title">Welcome, {st.session_state.user_name}</div>
      <div class="ap-hero-sub">Agent Code · {agent}</div>
      <div class="m-grid">
        <div class="m-card"><div class="m-num" style="color:#1e2d45">{s['total']}</div><div class="m-lbl">Total</div></div>
        <div class="m-card"><div class="m-num" style="color:#15803d">{s['done']}</div><div class="m-lbl">Done</div></div>
        <div class="m-card"><div class="m-num" style="color:#b45309">{s['recall']}</div><div class="m-lbl">Recall</div></div>
        <div class="m-card"><div class="m-num" style="color:#b91c1c">{s['closed']}</div><div class="m-lbl">Closed</div></div>
        <div class="m-card"><div class="m-num" style="color:#6d28d9">{s['fb'].get('out of area',0)}</div><div class="m-lbl">Out of Area</div></div>
        <div class="m-card"><div class="m-num" style="color:#475569">{s['ni']}</div><div class="m-lbl">Not Int.</div></div>
      </div>
    </div>""", unsafe_allow_html=True)

    st.markdown('<div class="sec-title">KPI Bars</div>', unsafe_allow_html=True)
    st.markdown(f"""
    <div class="kpi-grid">
      <div class="kpi-card">
        <div class="kpi-ttl">Done Rate</div>
        <div class="kpi-val" style="color:#15803d">{pct(s['done'],s['total'])}%</div>
        <div class="kpi-sub">{s['done']} of {s['total']}</div>
        {bar_html(s['done'],s['total'],'#15803d')}
      </div>
      <div class="kpi-card">
        <div class="kpi-ttl">Recall Rate</div>
        <div class="kpi-val" style="color:#b45309">{pct(s['recall'],s['total'])}%</div>
        <div class="kpi-sub">{s['recall']} clients</div>
        {bar_html(s['recall'],s['total'],'#b45309')}
      </div>
      <div class="kpi-card">
        <div class="kpi-ttl">Closed Rate</div>
        <div class="kpi-val" style="color:#b91c1c">{pct(s['closed'],s['total'])}%</div>
        <div class="kpi-sub">{s['closed']} clients</div>
        {bar_html(s['closed'],s['total'],'#b91c1c')}
      </div>
      <div class="kpi-card">
        <div class="kpi-ttl">Not Interested</div>
        <div class="kpi-val" style="color:#475569">{pct(s['ni'],s['total'])}%</div>
        <div class="kpi-sub">{s['ni']} clients</div>
        {bar_html(s['ni'],s['total'],'#475569')}
      </div>
    </div>""", unsafe_allow_html=True)

    if mine:
        st.markdown('<div class="sec-title">Feedback Breakdown</div>', unsafe_allow_html=True)
        _render_fb_cards(mine)
        st.markdown('<div class="sec-title">Data Source Breakdown</div>', unsafe_allow_html=True)
        _render_ds_cards(mine)
        st.markdown('<div class="sec-title">Export</div>', unsafe_allow_html=True)
        c1,_ = st.columns([1,3])
        with c1:
            st.download_button("⬇️ Download HTML Report",
                data=_build_html_report(mine, agent_label=agent),
                file_name=f"apply_agent{agent}_{datetime.now():%Y%m%d}.html", mime="text/html")

# ══════════════════════════════════════════════════════════════════════
#  PAGE: SALES — MY CLIENTS
# ══════════════════════════════════════════════════════════════════════
def page_my_clients():
    records = _load_records()
    agent   = str(st.session_state.agent_code).strip()
    mine    = [r for r in records if str(r.get("Agent Code","")).strip() == agent]

    st.markdown('<div class="sec-title">My Clients</div>', unsafe_allow_html=True)
    if not mine: st.info("No clients assigned yet."); return

    fc1,fc2 = st.columns(2)
    with fc1: srch = st.text_input("🔍 Search name / mobile")
    with fc2: filt = st.selectbox("Filter by Feedback", ["All"] + FEEDBACK_OPTIONS)

    show = mine[:]
    if srch:
        s    = srch.lower().lstrip("0")
        show = [r for r in show if s in str(r.get("Customer Name","")).lower()
                                or s in str(r.get("Mobile","")).lstrip("0")]
    if filt != "All":
        show = [r for r in show if str(r.get("Feedback (Sales)","")).strip() == filt]

    st.caption(f"Showing **{len(show)}** of **{len(mine)}** clients")

    for pos, rec in enumerate(show):
        mob      = str(rec.get("Mobile","")).strip()
        fb       = str(rec.get("Feedback (Sales)","")).strip()
        ds       = str(rec.get("Data Source Feedback","")).strip()
        note     = str(rec.get("Notes",""))
        ds_extra = str(rec.get("DS Extra",""))
        fb_col   = FB_COLORS.get(fb, "#64748b")
        ds_col   = DS_COLORS.get(ds, "#64748b")

        mp_html = ""
        if fb == "Member With other Phone":
            mp = str(rec.get("Member Phone",""))
            if mp and mp not in ("None","nan",""):
                mp_html = f'&nbsp;|&nbsp; 📱 <span style="color:#5b21b6;font-weight:600">{mp}</span>'
        ex_html = ""
        if ds in DS_NEEDS_EXTRA and ds_extra and ds_extra not in ("None","nan",""):
            ex_html = f'&nbsp;|&nbsp; <span style="color:#0369a1">{ds_extra}</span>'

        st.markdown(f"""
        <div class="client-card">
          <div style="display:flex;justify-content:space-between;align-items:center;flex-wrap:wrap;gap:6px">
            <span class="c-name">{rec.get('Customer Name','')}</span>
            <div>
              <span class="badge" style="background:{FB_BG.get(fb,'#f1f5f9')};color:{fb_col};border:1px solid {fb_col}30">{fb}</span>
              <span class="badge" style="background:{DS_BG.get(ds,'#f1f5f9')};color:{ds_col};border:1px solid {ds_col}30">{ds}</span>
            </div>
          </div>
          <div class="c-meta">
            📱 {mob} &nbsp;|&nbsp; 📅 {rec.get('Assign Data','')}
            {mp_html}{ex_html}
            {f'&nbsp;|&nbsp; 📝 {note}' if note and note not in ("None","nan","") else ''}
          </div>
        </div>""", unsafe_allow_html=True)

        first_word = (rec.get("Customer Name","Client") or "Client").split()[0]
        with st.expander(f"✏️ Update — {first_word}"):
            cc1, cc2 = st.columns(2)
            with cc1:
                st.text_input("Name",   value=rec.get("Customer Name",""), disabled=True, key=f"n_{agent}_{pos}")
                st.text_input("Mobile", value=mob, disabled=True, key=f"m_{agent}_{pos}")
                fidx   = FEEDBACK_OPTIONS.index(fb) if fb in FEEDBACK_OPTIONS else 0
                new_fb = st.selectbox("Feedback", FEEDBACK_OPTIONS, index=fidx, key=f"fb_{agent}_{pos}")
                new_member_phone = ""
                if new_fb == "Member With other Phone":
                    cur_mp = str(rec.get("Member Phone",""))
                    st.markdown('<div class="extra-box">', unsafe_allow_html=True)
                    new_member_phone = st.text_input(
                        "📱 Member's other phone number",
                        value=cur_mp if cur_mp not in ("None","nan","") else "",
                        placeholder="01XXXXXXXXX", key=f"mp_{agent}_{pos}")
                    st.markdown("</div>", unsafe_allow_html=True)
            with cc2:
                cur_ds = ds if ds in DS_OPTIONS else DS_OPTIONS[0]
                new_ds = st.selectbox("Data Source", DS_OPTIONS,
                                      index=DS_OPTIONS.index(cur_ds), key=f"ds_{agent}_{pos}")
                new_ds_extra = ""
                cur_ex = ds_extra if ds_extra not in ("None","nan","") else ""
                if new_ds == "recommendation":
                    st.markdown('<div class="extra-box">', unsafe_allow_html=True)
                    new_ds_extra = st.text_input("👤 Recommender name & phone",
                        value=cur_ex, placeholder="Ahmed Ali — 01XXXXXXXXX", key=f"dsx_{agent}_{pos}")
                    st.markdown("</div>", unsafe_allow_html=True)
                elif new_ds == "Facebook":
                    st.markdown('<div class="extra-box">', unsafe_allow_html=True)
                    new_ds_extra = st.text_input("📘 Facebook details",
                        value=cur_ex, placeholder="Campaign / post / page…", key=f"dsx_{agent}_{pos}")
                    st.markdown("</div>", unsafe_allow_html=True)
                elif new_ds == "other":
                    st.markdown('<div class="extra-box">', unsafe_allow_html=True)
                    new_ds_extra = st.text_input("📝 Other — please specify",
                        value=cur_ex, placeholder="Describe the source…", key=f"dsx_{agent}_{pos}")
                    st.markdown("</div>", unsafe_allow_html=True)
                new_nt = st.text_area("Notes",
                    value=note if note not in ("None","nan","") else "",
                    height=80, key=f"nt_{agent}_{pos}")

            if st.button("💾 Save", key=f"save_{agent}_{pos}", type="primary"):
                all_recs = _load_records()
                for r2 in all_recs:
                    if (str(r2.get("Agent Code","")).strip() == agent and
                            str(r2.get("Mobile","")).strip() == mob):
                        r2["Feedback (Sales)"]     = new_fb
                        r2["Data Source Feedback"] = new_ds
                        r2["DS Extra"]             = new_ds_extra
                        r2["Notes"]                = new_nt
                        if new_fb == "Member With other Phone":
                            r2["Member Phone"] = new_member_phone
                        break
                _save_records(all_recs)
                st.success("✅ Saved!")
                st.rerun()

    st.markdown('<div class="sec-title">Download</div>', unsafe_allow_html=True)
    c1,_ = st.columns([1,3])
    with c1:
        st.download_button("⬇️ Download My Data (Excel)",
            data=_to_excel_bytes(mine),
            file_name=f"agent_{agent}_{datetime.now():%Y%m%d}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ══════════════════════════════════════════════════════════════════════
#  EXCEL DASHBOARD
# ══════════════════════════════════════════════════════════════════════
def _build_excel_dashboard(records: list) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    wb.remove(wb.active)

    # ── palette (no # prefix for openpyxl) ──
    H_BLUE  = "1E40AF"
    H_LBLUE = "DBEAFE"
    H_TOTAL = "EFF6FF"
    H_WHITE = "FFFFFF"
    H_GREY  = "F8FAFC"
    T_WHITE = "FFFFFF"
    T_DARK  = "1E2D45"
    T_GREY  = "64748B"

    THIN  = Side(style="thin",   color="D0DAE8")
    bdr   = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    def _s(ws, r, c, val, bg=H_WHITE, fg=T_DARK, bold=False, sz=10,
           halign="center", fmt=None, wrap=False):
        """Write a single (non-merged) cell safely."""
        cell = ws.cell(row=r, column=c)
        cell.value     = val
        cell.font      = Font(name="Calibri", bold=bold, size=sz, color=fg)
        cell.fill      = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal=halign, vertical="center",
                                    wrap_text=wrap)
        cell.border    = bdr
        if fmt:
            cell.number_format = fmt
        return cell

    def _hdr(ws, r, c, val, bg=H_BLUE, fg=T_WHITE, sz=10, wrap=False):
        return _s(ws, r, c, val, bg=bg, fg=fg, bold=True, sz=sz, wrap=wrap)

    s      = _get_stats(records)
    now    = datetime.now().strftime("%Y-%m-%d %H:%M")
    total  = s["total"] or 1
    agents = sorted({str(r.get("Agent Code","")).strip()
                     for r in records if r.get("Agent Code")})

    # ══════════════════════════════════════════════════
    # SHEET 1 — OVERVIEW
    # ══════════════════════════════════════════════════
    ws1 = wb.create_sheet("Overview")
    ws1.sheet_view.showGridLines = False

    # column widths
    for ci, w in [(1,22),(2,8),(3,8),(4,8)]:
        ws1.column_dimensions[get_column_letter(ci)].width = w

    # ── Row 1: title (plain, no merge)
    ws1.row_dimensions[1].height = 34
    tc = ws1.cell(1,1, "APPLY — Sales Dashboard")
    tc.font      = Font(name="Calibri", bold=True, size=16, color=H_BLUE)
    tc.alignment = Alignment(horizontal="left", vertical="center")
    ws1.cell(1,4, now).font = Font(name="Calibri", size=9, color="94A3B8")

    # ── Row 2: blank spacer
    ws1.row_dimensions[2].height = 8

    # ── Rows 3-4: KPI block
    kpis = [
        ("Total",   s["total"],                  H_LBLUE, T_DARK,  None),
        ("Done",    s["done"],                   "DCFCE7","15803D",None),
        ("Done %",  pct(s["done"],s["total"])/100,"DCFCE7","15803D","0%"),
        ("Recall",  s["recall"],                 "FEF3C7","B45309",None),
        ("Closed",  s["closed"],                 "FEE2E2","B91C1C",None),
        ("Not Int.",s["ni"],                     "F1F5F9","475569",None),
    ]
    ws1.row_dimensions[3].height = 22
    ws1.row_dimensions[4].height = 30
    # put KPIs side-by-side in col A (label) + col B-G (values horizontally)
    # Actually put each KPI in its own column pair would need more cols — 
    # let's just use one column layout: col A = label, col B = value
    for ki, (lbl, val, bg, fg, fmt) in enumerate(kpis):
        row_l = 3 + ki*2
        row_v = 4 + ki*2
        ws1.row_dimensions[row_l].height = 16
        ws1.row_dimensions[row_v].height = 26
        _hdr(ws1, row_l, 1, lbl, bg=bg, fg=fg, sz=9)
        c = _s(ws1, row_v, 1, val if not fmt else val,
               bg=bg, fg=fg, bold=True, sz=16, fmt=fmt)

    # ── Feedback section (starts after KPIs)
    fb_row = 3 + len(kpis)*2 + 1
    ws1.row_dimensions[fb_row].height = 8   # spacer
    fb_row += 1

    _hdr(ws1, fb_row,   1, "Feedback",  sz=10)
    _hdr(ws1, fb_row,   2, "Count",     sz=10)
    _hdr(ws1, fb_row,   3, "%",         sz=10)
    fb_row += 1

    for fb in FEEDBACK_OPTIONS:
        cnt = s["fb"].get(fb, 0)
        bg  = FB_BG.get(fb,  "#F8FAFC").replace("#","")
        fg  = FB_COLORS.get(fb, "#1E2D45").replace("#","")
        ws1.row_dimensions[fb_row].height = 18
        _s(ws1, fb_row, 1, fb,       bg=bg, fg=fg, halign="left")
        _s(ws1, fb_row, 2, cnt,      bg=bg, fg=fg, bold=True)
        _s(ws1, fb_row, 3, cnt/total,bg=bg, fg=fg, fmt="0.0%")
        fb_row += 1

    # ── Data Source section
    fb_row += 1  # spacer
    _hdr(ws1, fb_row, 1, "Data Source", sz=10)
    _hdr(ws1, fb_row, 2, "Count",       sz=10)
    _hdr(ws1, fb_row, 3, "%",           sz=10)
    fb_row += 1

    for ds in DS_OPTIONS:
        cnt = s["ds"].get(ds, 0)
        bg  = DS_BG.get(ds,  "#F8FAFC").replace("#","")
        fg  = DS_COLORS.get(ds, "#1E2D45").replace("#","")
        ws1.row_dimensions[fb_row].height = 18
        _s(ws1, fb_row, 1, ds,       bg=bg, fg=fg, halign="left")
        _s(ws1, fb_row, 2, cnt,      bg=bg, fg=fg, bold=True)
        _s(ws1, fb_row, 3, cnt/total,bg=bg, fg=fg, fmt="0.0%")
        fb_row += 1

    # ══════════════════════════════════════════════════
    # SHEET 2 — AGENT PERFORMANCE
    # ══════════════════════════════════════════════════
    ws2 = wb.create_sheet("Agent Performance")
    ws2.sheet_view.showGridLines = False

    cols = ["Agent", "Total"] + FEEDBACK_OPTIONS + ["Conv %"]
    for ci, col_name in enumerate(cols, start=1):
        ws2.column_dimensions[get_column_letter(ci)].width = max(14, len(col_name)+3)
        bg = H_BLUE
        fg = T_WHITE
        if col_name in FB_BG:
            bg = FB_BG[col_name].replace("#","")
            fg = FB_COLORS.get(col_name, T_DARK).replace("#","")
        _hdr(ws2, 1, ci, col_name, bg=bg, fg=fg, sz=10, wrap=True)
    ws2.row_dimensions[1].height = 42

    ag_rows = []
    for ac in agents:
        mine  = [r for r in records if str(r.get("Agent Code","")).strip() == ac]
        ag_s  = _get_stats(mine)
        ag_rows.append((ac, ag_s))
    ag_rows.sort(key=lambda x: x[1]["done"], reverse=True)

    for ri, (ac, ag_s) in enumerate(ag_rows, start=2):
        done_pct = ag_s["done"] / (ag_s["total"] or 1)
        ws2.row_dimensions[ri].height = 20
        ci = 1
        _s(ws2, ri, ci, ac, bold=True, halign="left"); ci += 1
        _s(ws2, ri, ci, ag_s["total"], bg=H_LBLUE, fg=T_DARK, bold=True); ci += 1
        for fb in FEEDBACK_OPTIONS:
            cnt = ag_s["fb"].get(fb, 0)
            bg  = FB_BG.get(fb,  "#FFFFFF").replace("#","")
            fg  = FB_COLORS.get(fb, T_DARK).replace("#","")
            _s(ws2, ri, ci, cnt if cnt > 0 else "",
               bg=bg if cnt > 0 else H_WHITE, fg=fg, bold=(cnt > 0))
            ci += 1
        _s(ws2, ri, ci, done_pct,
           bg="DCFCE7" if done_pct > 0 else H_WHITE,
           fg="15803D", bold=True, fmt="0%")

    # Totals row
    tr = len(ag_rows) + 2
    ws2.row_dimensions[tr].height = 24
    ci = 1
    _s(ws2, tr, ci, "TOTAL", bg=H_LBLUE, fg=T_DARK, bold=True, halign="left"); ci += 1
    _s(ws2, tr, ci, s["total"], bg=H_LBLUE, fg=T_DARK, bold=True, sz=11); ci += 1
    for fb in FEEDBACK_OPTIONS:
        cnt = s["fb"].get(fb, 0)
        bg  = FB_BG.get(fb, "#EFF6FF").replace("#","")
        fg  = FB_COLORS.get(fb, T_DARK).replace("#","")
        _s(ws2, tr, ci, cnt, bg=bg, fg=fg, bold=True, sz=11); ci += 1
    _s(ws2, tr, ci, pct(s["done"], s["total"])/100,
       bg="DCFCE7", fg="15803D", bold=True, sz=11, fmt="0%")

    # ══════════════════════════════════════════════════
    # SHEET 3 — RAW DATA
    # ══════════════════════════════════════════════════
    ws3 = wb.create_sheet("Raw Data")
    ws3.sheet_view.showGridLines = False

    if records:
        df = pd.DataFrame(records).copy()
        if "Mobile" in df.columns:
            df["Mobile"] = df["Mobile"].apply(
                lambda x: str(int(float(str(x)))).zfill(11)
                if str(x).strip() not in ("","None","nan") else "")
        col_names = list(df.columns)
        for ci, cn in enumerate(col_names, start=1):
            ws3.column_dimensions[get_column_letter(ci)].width = max(14, len(str(cn))+4)
            _hdr(ws3, 1, ci, cn, sz=10)
        ws3.row_dimensions[1].height = 22
        for ri, row in df.iterrows():
            ws3.row_dimensions[ri+2].height = 18
            for ci, cn in enumerate(col_names, start=1):
                v = row[cn]
                if cn == "Feedback (Sales)":
                    fb_val = str(v).strip()
                    bg = FB_BG.get(fb_val, "#FFFFFF").replace("#","")
                    fg = FB_COLORS.get(fb_val, T_DARK).replace("#","")
                    _s(ws3, ri+2, ci, v, bg=bg, fg=fg)
                elif cn == "Mobile":
                    c2 = ws3.cell(row=ri+2, column=ci)
                    c2.value          = str(v) if v else ""
                    c2.font           = Font(name="Calibri", size=10)
                    c2.fill           = PatternFill("solid", fgColor=H_WHITE)
                    c2.alignment      = Alignment(horizontal="center", vertical="center")
                    c2.border         = bdr
                    c2.number_format  = "@"
                else:
                    _s(ws3, ri+2, ci, v, halign="left")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

# ══════════════════════════════════════════════════════════════════════
#  HTML REPORT# ══════════════════════════════════════════════════════════════════════
#  HTML REPORT
# ══════════════════════════════════════════════════════════════════════
def _build_html_report(records: list, agent_label: str = None, title: str = None) -> str:
    s     = _get_stats(records)
    total = s["total"]
    conv  = pct(s["done"], total)
    now   = datetime.now().strftime("%Y-%m-%d %H:%M")

    fb_cnt   = Counter(str(r.get("Feedback (Sales)","")).strip() for r in records)
    fb_items = sorted(fb_cnt.items(), key=lambda x: x[1], reverse=True)
    mx_fb    = fb_items[0][1] if fb_items else 1

    fb_cards = ""
    for lb, v in fb_items:
        if not lb or lb in ("None","nan",""): continue
        fg = FB_COLORS.get(lb, "#475569"); bg = FB_BG.get(lb, "#f8fafc")
        fb_cards += (
            f'<div style="background:{bg};border:1px solid {fg}25;border-left:4px solid {fg};'
            f'border-radius:12px;padding:20px 14px;text-align:center">'
            f'<div style="font-family:Syne,sans-serif;font-size:2.6rem;font-weight:900;color:{fg};line-height:1">{v}</div>'
            f'<div style="font-size:13px;color:{fg};margin-top:8px;font-weight:700;text-transform:uppercase;letter-spacing:.5px">{lb}</div>'
            f'<div style="font-size:14px;color:#64748b;margin-top:4px;font-weight:600">{pct(v,total)}%</div>'
            f'</div>'
        )

    fb_bars = ""
    for lb, v in fb_items:
        if not lb or lb in ("None","nan",""): continue
        fg = FB_COLORS.get(lb, "#475569")
        fb_bars += (
            f'<div style="display:flex;align-items:center;gap:14px;margin:12px 0">'
            f'<span style="width:185px;font-size:15px;color:#334155;flex-shrink:0;font-weight:600">{lb}</span>'
            f'<div style="flex:1;background:#f1f5f9;border-radius:6px;height:24px;overflow:hidden">'
            f'<div style="background:{fg};width:{int(v/mx_fb*100)}%;height:24px;border-radius:6px"></div></div>'
            f'<span style="font-size:17px;font-weight:900;color:{fg};min-width:32px">{v}</span>'
            f'<span style="font-size:14px;color:#94a3b8;min-width:42px">{pct(v,total)}%</span>'
            f'</div>'
        )

    ds_cnt   = Counter(str(r.get("Data Source Feedback","")).strip() for r in records)
    ds_items = sorted(ds_cnt.items(), key=lambda x: x[1], reverse=True)
    ds_cards = ""
    for ds, cnt in ds_items:
        if not ds or ds in ("None","nan",""): continue
        fg = DS_COLORS.get(ds, "#475569"); bg = DS_BG.get(ds, "#f8fafc")
        ds_cards += (
            f'<div style="background:{bg};border:1px solid {fg}25;border-left:4px solid {fg};'
            f'border-radius:12px;padding:20px 14px;text-align:center">'
            f'<div style="font-family:Syne,sans-serif;font-size:2.6rem;font-weight:900;color:{fg};line-height:1">{cnt}</div>'
            f'<div style="font-size:13px;color:{fg};margin-top:8px;font-weight:700;text-transform:uppercase;letter-spacing:.5px">{ds}</div>'
            f'<div style="font-size:14px;color:#64748b;margin-top:4px;font-weight:600">{pct(cnt,total)}%</div>'
            f'</div>'
        )

    agent_section = ""
    if not agent_label:
        agents  = sorted({str(r.get("Agent Code","")).strip() for r in records if r.get("Agent Code")})
        ag_data = []
        for ac in agents:
            mine  = [r for r in records if str(r.get("Agent Code","")).strip() == ac]
            ag_s  = _get_stats(mine)
            ag_data.append((ac, ag_s))
        ag_data.sort(key=lambda x: x[1]["done"], reverse=True)
        rows = ""
        for rank, (ac, ag_s) in enumerate(ag_data):
            medal    = ["🥇","🥈","🥉"][rank] if rank < 3 else ""
            done_pct = pct(ag_s["done"], ag_s["total"])
            rows += (
                f"<tr>"
                f"<td style='font-weight:700;color:#1e2d45;font-size:15px'>{medal} {ac}</td>"
                f"<td style='color:#334155;font-size:15px'>{ag_s['total']}</td>"
                f"<td style='color:#15803d;font-weight:800;font-size:18px'>{ag_s['done']}</td>"
                f"<td style='color:#b45309;font-size:15px'>{ag_s['recall']}</td>"
                f"<td style='color:#b91c1c;font-size:15px'>{ag_s['closed']}</td>"
                f"<td><div style='display:flex;align-items:center;gap:8px'>"
                f"<div style='flex:1;background:#f1f5f9;border-radius:4px;height:12px;overflow:hidden'>"
                f"<div style='background:#15803d;width:{done_pct}%;height:12px;border-radius:4px'></div></div>"
                f"<span style='color:#15803d;font-weight:800;font-size:16px;min-width:42px'>{done_pct}%</span>"
                f"</div></td>"
                f"</tr>"
            )
        agent_section = f"""
        <div class="section">
          <h2>Agent Performance — Sorted by Done ↓</h2>
          <table>
            <tr><th>Agent</th><th>Total</th><th>Done</th><th>Recall</th><th>Closed</th><th>Conv %</th></tr>
            {rows}
          </table>
        </div>"""

    if title:
        subtitle = title
    elif agent_label:
        subtitle = f"Agent {agent_label}"
    else:
        subtitle = "Full Team Report"

    return f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Apply — {subtitle}</title>
<style>
@import url('https://fonts.googleapis.com/css2?family=Syne:wght@700;900&family=DM+Sans:wght@400;500;600;700&display=swap');
*{{box-sizing:border-box;margin:0;padding:0}}
body{{font-family:'DM Sans',sans-serif;background:#f6f8fc;color:#1e2d45;padding:44px 52px;line-height:1.6;font-size:15px}}
h1{{font-family:'Syne',sans-serif;font-size:2.6rem;font-weight:900;
    background:linear-gradient(135deg,#1d4ed8,#0891b2);
    -webkit-background-clip:text;-webkit-text-fill-color:transparent;margin-bottom:6px}}
.sub{{color:#64748b;font-size:15px;margin-bottom:36px;font-weight:500}}
.kpis{{display:grid;grid-template-columns:repeat(4,1fr);gap:18px;margin-bottom:30px}}
.kpi{{background:#ffffff;border:1px solid #e4eaf3;border-radius:16px;padding:26px 22px;text-align:center;
      box-shadow:0 2px 10px rgba(0,0,0,.05)}}
.kpi .v{{font-family:'Syne',sans-serif;font-size:3rem;font-weight:900;color:#1e2d45;margin-bottom:6px;line-height:1}}
.kpi .l{{color:#94a3b8;font-size:13px;text-transform:uppercase;letter-spacing:.8px;font-weight:700}}
.section{{background:#ffffff;border:1px solid #e4eaf3;border-radius:16px;padding:28px 32px;
          margin-bottom:24px;box-shadow:0 2px 10px rgba(0,0,0,.04)}}
h2{{font-family:'Syne',sans-serif;font-size:1.05rem;font-weight:800;
    color:#1e40af;text-transform:uppercase;letter-spacing:1px;margin-bottom:22px;
    padding-bottom:12px;border-bottom:2px solid #dbeafe}}
.cards-grid{{display:grid;grid-template-columns:repeat(auto-fit,minmax(130px,1fr));gap:14px;margin-bottom:24px}}
table{{width:100%;border-collapse:collapse}}
th{{text-align:left;padding:13px 18px;color:#94a3b8;font-size:12px;
    text-transform:uppercase;letter-spacing:.7px;border-bottom:2px solid #f1f5f9;font-weight:700}}
td{{padding:15px 18px;border-bottom:1px solid #f8fafc;font-size:15px;color:#334155;vertical-align:middle}}
tr:hover td{{background:#fafbff}}
.foot{{text-align:center;color:#cbd5e1;font-size:13px;margin-top:40px;padding-top:20px;border-top:1px solid #f1f5f9}}
@media print{{body{{background:#fff;padding:20px}}.section,.kpi{{box-shadow:none}}}}
</style>
</head>
<body>
<h1>APPLY</h1>
<div class="sub">{subtitle} &nbsp;·&nbsp; {now}</div>
<div class="kpis">
  <div class="kpi"><div class="v">{total}</div><div class="l">Total Clients</div></div>
  <div class="kpi"><div class="v" style="color:#15803d">{s['done']}</div>
    <div class="l">Done &nbsp;<span style="color:#15803d;font-size:16px;font-weight:800">{conv}%</span></div></div>
  <div class="kpi"><div class="v" style="color:#b45309">{s['recall']}</div><div class="l">Recall</div></div>
  <div class="kpi"><div class="v" style="color:#b91c1c">{s['closed']}</div><div class="l">Closed</div></div>
</div>
<div class="section">
  <h2>Feedback Breakdown</h2>
  <div class="cards-grid">{fb_cards}</div>
  {fb_bars}
</div>
<div class="section">
  <h2>Data Source</h2>
  <div class="cards-grid">{ds_cards}</div>
</div>
{agent_section}
<div class="foot">Apply Sales Management Portal &nbsp;·&nbsp; {now}</div>
</body>
</html>"""

# ══════════════════════════════════════════════════════════════════════
#  ROUTING
# ══════════════════════════════════════════════════════════════════════
if st.session_state.role == "admin":
    if   page == "Dashboard":       page_dashboard()
    elif page == "All Data":        page_all_data()
    elif page == "Upload Data":     page_upload()
    elif page == "Users":           page_users()
    elif page == "Monthly Report":  page_monthly_report()
else:
    if   page == "My Dashboard": page_my_dashboard()
    elif page == "My Clients":   page_my_clients()
